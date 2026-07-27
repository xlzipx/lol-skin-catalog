"""CSV and XLSX output."""

import csv
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

from . import i18n
from .theme import TIERS, TIER_HEX

IMAGE_WIDTH_PX = 200


def write_csv(skins, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([
            i18n.t("col_index"), i18n.t("col_champion"), i18n.t("col_skin"),
            i18n.t("col_rarity"), i18n.t("col_chromas"), i18n.t("col_skin_id"),
            i18n.t("col_splash_file"),
        ])
        for i, skin in enumerate(skins, 1):
            w.writerow([
                i, skin["champion"], skin["skin"], skin.get("rarity", ""),
                skin["chromas"] or "", skin["skinId"],
                os.path.basename(skin["file"]) if skin.get("file") else "",
            ])


def write_xlsx(skins, profile, tier_counts, path):
    wb = Workbook()
    ws = wb.active
    ws.title = i18n.t("sheet_skins")

    dark = PatternFill("solid", fgColor="0A1428")
    gold_font = Font(bold=True, color="C8AA6E", size=11)
    thin = Side(style="thin", color="1E2D44")

    headers = [
        i18n.t("col_index"), i18n.t("col_champion"), i18n.t("col_skin"),
        i18n.t("col_rarity"), i18n.t("col_chromas"), i18n.t("col_skin_id"),
        i18n.t("col_splash"),
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = gold_font
        cell.fill = dark
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    for width, letter in zip([5, 18, 34, 13, 9, 10, 34], "ABCDEFG"):
        ws.column_dimensions[letter].width = width

    for i, skin in enumerate(skins, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=skin["champion"])
        ws.cell(row=r, column=3, value=skin["skin"])
        rarity_cell = ws.cell(row=r, column=4, value=skin.get("rarity", "") or "—")
        ws.cell(row=r, column=5, value=skin["chromas"] or None)
        ws.cell(row=r, column=6, value=skin["skinId"])

        tier = skin.get("rarity", "")
        if tier:
            rarity_cell.fill = PatternFill("solid", fgColor=TIER_HEX[tier])
            rarity_cell.font = Font(bold=True, color="0A1428", size=10)
        else:
            rarity_cell.font = Font(color="8A8A8A")

        for c in (1, 4, 5, 6):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="center")
        for c in (2, 3):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = Border(bottom=thin)

        if skin.get("thumb") and os.path.exists(skin["thumb"]):
            with Image.open(skin["thumb"]) as im:
                ratio = im.height / im.width
            img = XLImage(skin["thumb"])
            img.width = IMAGE_WIDTH_PX
            img.height = round(IMAGE_WIDTH_PX * ratio)
            ws.add_image(img, f"{get_column_letter(7)}{r}")
            ws.row_dimensions[r].height = round(IMAGE_WIDTH_PX * ratio) * 0.78 + 6
        else:
            ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = f"A1:F{len(skins) + 1}"

    # ------------------------------------------------------------ summary --
    ws2 = wb.create_sheet(i18n.t("sheet_summary"))
    for letter, width in (("A", 24), ("B", 14), ("D", 22), ("E", 14)):
        ws2.column_dimensions[letter].width = width

    def heading(row, col, text):
        cell = ws2.cell(row=row, column=col, value=text)
        cell.font = gold_font
        cell.fill = dark

    heading(1, 1, i18n.t("xls_overview"))
    heading(1, 2, "")
    facts = [
        (i18n.t("xls_player"),
         f"{profile.get('gameName', '')} #{profile.get('tagLine', '')}".strip()),
        (i18n.t("xls_level"), profile.get("level")),
        (i18n.t("xls_skins_total"), len(skins)),
        (i18n.t("xls_champs_with_skin"), len({s["champion"] for s in skins})),
        (i18n.t("xls_champs_owned"), profile.get("championsOwned")),
        (i18n.t("xls_chromas_owned"), profile.get("chromasOwned")),
        (i18n.t("xls_export_date"), date.today().strftime("%d.%m.%Y")),
    ]
    for i, (key, value) in enumerate(facts, start=2):
        ws2.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)

    row = len(facts) + 3
    heading(row, 1, i18n.t("col_rarity"))
    heading(row, 2, i18n.t("col_count"))
    for tier in TIERS:
        row += 1
        cell = ws2.cell(row=row, column=1, value=tier)
        cell.fill = PatternFill("solid", fgColor=TIER_HEX[tier])
        cell.font = Font(bold=True, color="0A1428")
        ws2.cell(row=row, column=2, value=tier_counts.get(tier, 0))
    row += 1
    ws2.cell(row=row, column=1, value=i18n.t("xls_no_tier")).font = Font(color="8A8A8A")
    ws2.cell(row=row, column=2, value=tier_counts.get("", 0))

    heading(1, 4, i18n.t("col_champion"))
    heading(1, 5, i18n.t("col_count"))
    counts = {}
    for skin in skins:
        counts[skin["champion"]] = counts.get(skin["champion"], 0) + 1
    for i, champion in enumerate(sorted(counts, key=str.lower), start=2):
        ws2.cell(row=i, column=4, value=champion)
        ws2.cell(row=i, column=5, value=counts[champion])
    ws2.freeze_panes = "A2"

    wb.save(path)
